"""Generate styled multi-sheet Excel report for Kaamil Technology Sales.

Sheets: Summary, by_Category, by_Location, by_SalesRep, monthly_trend,
        price_tiers, ML_Model_Performance, ML_Features.
Uses the xlsx skill's conventions: openpyxl styling, tables, conditional
formatting, and a basic bar chart on the Summary sheet.
"""
import json
import os
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import PROCESSED_FILE, OUTPUT_DIR, EXCEL_REPORT_PATH

# Colors
NAVY = "001E293B"
BLUE = "002563EB"
GREEN = "0010B981"
AMBER = "00F59E0B"
PURPLE = "00A855F7"
WHITE = "00FFFFFF"
LIGHT = "00F1F5F9"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Calibri", bold=True, color="00FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color="00FFFFFF", size=14)
SUBTITLE_FONT = Font(name="Calibri", bold=True, color="001E293B", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
THIN = Side(style="thin", color="00CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def style_header_row(ws, row, ncols, fill=HEADER_FILL):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER


def style_data_area(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL_FONT
            cell.border = BORDER
            cell.alignment = LEFT if c == 1 else RIGHT


def add_table(ws, ref, name):
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                                        showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)


def build_report():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    df = pd.read_csv(PROCESSED_FILE)
    df["Transaction_Date"] = pd.to_datetime(df["Transaction_Date"], errors="coerce")
    profile = json.load(open(OUTPUT_DIR / "eda_profile.json"))
    results = json.load(open(OUTPUT_DIR / "results.json"))

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ============================================================
    # SHEET 1: Summary KPIs
    # ============================================================
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "Kaamil Technology Sales — Analytics Report"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Period: {profile['date_range']['start']} → {profile['date_range']['end']}  ·  {profile['raw_rows']} transactions  ·  {len(profile['location_counts'])} cities"
    ws["A2"].font = Font(name="Calibri", color="0094A3B8", size=11)
    ws["A2"].alignment = CENTER

    # KPI cards as a small table
    kpi_headers = ["Metric", "Value"]
    kpi_data = [
        ("Total Revenue", f"${profile['total_revenue']:,.0f}"),
        ("Total Profit", f"${profile['total_profit']:,.0f}"),
        ("Avg Order Value", f"${profile['avg_order_value']:,.0f}"),
        ("Avg Profit Margin", f"{profile['avg_profit_margin']*100:.1f}%"),
        ("Best ML Model", f"{results['best_model']} (R²={results['best_r2_score']:.4f})"),
        ("Categories", ", ".join(profile['category_counts'].keys())),
    ]
    ws.append([])
    ws.append(kpi_headers)
    for row in kpi_data:
        ws.append(list(row))
    style_header_row(ws, 4, 2)
    style_data_area(ws, 5, 5 + len(kpi_data), 2)
    # Accent the values
    for r in range(5, 5 + len(kpi_data)):
        ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True, color="002563EB", size=12)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 34
    add_table(ws, f"A4:B{4+len(kpi_data)}", "KPI_Table")

    # Monthly trend table
    start_row = 5 + len(kpi_data) + 2
    ws.merge_cells(f"A{start_row}:D{start_row}")
    ws[f"A{start_row}"] = "Monthly Revenue Trend"
    ws[f"A{start_row}"].font = SUBTITLE_FONT
    ws[f"A{start_row}"].fill = PatternFill("solid", fgColor=LIGHT)
    start_row += 1
    ws.append([])
    ws.append(["Month", "Revenue ($)", "MoM Change (%)", "Profit Margin (%)"])
    style_header_row(ws, start_row, 4)
    month_names = {6: "Jun 2026", 7: "Jul 2026", 8: "Aug 2026"}
    prev = None
    m = start_row + 1
    for month in sorted([int(k) for k in profile["monthly_revenue"].keys()]):
        rev = float(profile["monthly_revenue"][str(month)])
        mom = "" if prev is None else f"{(rev / prev - 1) * 100:.1f}%"
        margin = f"{profile['avg_profit_margin']*100:.1f}%"
        ws.append([month_names.get(month, str(month)), round(rev, 2), mom, margin])
        prev = rev
        m += 1
    style_data_area(ws, start_row + 1, m - 1, 4)
    # Number format
    for r in range(start_row + 1, m):
        ws.cell(row=r, column=2).number_format = '$#,##0'
    add_table(ws, f"A{start_row}:D{m-1}", "Monthly_Table")

    # Bar chart for monthly trend
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Revenue"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Month"
    chart.style = 10
    data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=m - 1)
    cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=m - 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    chart.width = 14
    chart.height = 8
    ws.add_chart(chart, f"F{start_row}")

    for col in "ABCDEF":
        ws.column_dimensions[col].width = 16

    # ============================================================
    # SHEET 2: by Category
    # ============================================================
    ws = wb.create_sheet("by_Category")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "Performance by Product Category"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    cat_stats = df.groupby("Category").agg(
        Transactions=("Transaction_ID", "count"),
        Total_Revenue=("Total_Revenue_USD", "sum"),
        Total_Profit=("Profit_USD", "sum"),
        Avg_Revenue=("Total_Revenue_USD", "mean"),
        Avg_Profit_Margin=("Profit_Margin", "mean"),
        Avg_Quantity=("Quantity", "mean"),
    ).reset_index()
    cat_stats["Avg_Profit_Margin"] = cat_stats["Avg_Profit_Margin"].apply(lambda x: f"{x*100:.1f}%")
    cat_stats["Total_Revenue"] = cat_stats["Total_Revenue"].round(2)
    cat_stats["Total_Profit"] = cat_stats["Total_Profit"].round(2)
    cat_stats["Avg_Revenue"] = cat_stats["Avg_Revenue"].round(2)
    cat_stats = cat_stats.sort_values("Total_Revenue", ascending=False)

    cols = ["Category", "Transactions", "Total_Revenue", "Total_Profit", "Avg_Revenue", "Avg_Profit_Margin"]
    ws.append([])
    ws.append(cols)
    style_header_row(ws, 3, len(cols))
    for i, row in cat_stats.iterrows():
        ws.append(list(row))
    style_data_area(ws, 4, 3 + len(cat_stats), len(cols))
    for r in range(4, 4 + len(cat_stats)):
        ws.cell(row=r, column=3).number_format = '$#,##0'
        ws.cell(row=r, column=4).number_format = '$#,##0'
        ws.cell(row=r, column=5).number_format = '$#,##0'
    add_table(ws, f"A3:F{3+len(cat_stats)}", "Category_Table")

    # Pie chart of revenue by category
    chart = PieChart()
    chart.title = "Revenue by Category"
    data_ref = Reference(ws, min_col=3, min_row=3, max_row=3 + len(cat_stats))
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=3 + len(cat_stats))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 14
    chart.height = 8
    ws.add_chart(chart, "H3")

    for col_letter in "ABCDEF":
        ws.column_dimensions[col_letter].width = 18

    # ============================================================
    # SHEET 3: by Location
    # ============================================================
    ws = wb.create_sheet("by_Location")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    ws["A1"] = "Performance by City (Location)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    loc_stats = df.groupby("Location").agg(
        Transactions=("Transaction_ID", "count"),
        Total_Revenue=("Total_Revenue_USD", "sum"),
        Total_Profit=("Profit_USD", "sum"),
        Avg_Revenue=("Total_Revenue_USD", "mean"),
        Avg_Profit_Margin=("Profit_Margin", "mean"),
    ).reset_index()
    loc_stats["Avg_Profit_Margin"] = loc_stats["Avg_Profit_Margin"].apply(lambda x: f"{x*100:.1f}%")
    loc_stats["Total_Revenue"] = loc_stats["Total_Revenue"].round(2)
    loc_stats["Total_Profit"] = loc_stats["Total_Profit"].round(2)
    loc_stats["Avg_Revenue"] = loc_stats["Avg_Revenue"].round(2)
    loc_stats = loc_stats.sort_values("Total_Revenue", ascending=False)

    cols = ["Location", "Transactions", "Total_Revenue", "Total_Profit", "Avg_Revenue"]
    ws.append([])
    ws.append(cols)
    style_header_row(ws, 3, len(cols))
    for i, row in loc_stats.iterrows():
        ws.append(list(row))
    style_data_area(ws, 4, 3 + len(loc_stats), len(cols))
    for r in range(4, 4 + len(loc_stats)):
        ws.cell(row=r, column=3).number_format = '$#,##0'
        ws.cell(row=r, column=4).number_format = '$#,##0'
    add_table(ws, f"A3:E{3+len(loc_stats)}", "Location_Table")

    # Conditional colour scale on revenue
    ws.conditional_formatting.add(
        f"C4:C{3+len(loc_stats)}",
        ColorScaleRule(
            start_type="min", start_color="001E293B",
            mid_type="percentile", mid_value=50, mid_color="002563EB",
            end_type="max", end_color="0010B981",
        ),
    )

    for col_letter in "ABCDE":
        ws.column_dimensions[col_letter].width = 16

    # ============================================================
    # SHEET 4: by Sales Rep
    # ============================================================
    ws = wb.create_sheet("by_SalesRep")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    ws["A1"] = "Performance by Sales Representative"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    rep_stats = df.groupby("Sales_Rep").agg(
        Transactions=("Transaction_ID", "count"),
        Total_Revenue=("Total_Revenue_USD", "sum"),
        Total_Profit=("Profit_USD", "sum"),
        Avg_Revenue=("Total_Revenue_USD", "mean"),
        Avg_Profit_Margin=("Profit_Margin", "mean"),
    ).reset_index()
    rep_stats["Avg_Profit_Margin"] = rep_stats["Avg_Profit_Margin"].apply(lambda x: f"{x*100:.1f}%")
    rep_stats["Total_Revenue"] = rep_stats["Total_Revenue"].round(2)
    rep_stats["Total_Profit"] = rep_stats["Total_Profit"].round(2)
    rep_stats["Avg_Revenue"] = rep_stats["Avg_Revenue"].round(2)
    rep_stats = rep_stats.sort_values("Total_Revenue", ascending=False)

    cols = ["Sales_Rep", "Transactions", "Total_Revenue", "Total_Profit", "Avg_Revenue"]
    ws.append([])
    ws.append(cols)
    style_header_row(ws, 3, len(cols))
    for i, row in rep_stats.iterrows():
        ws.append(list(row))
    style_data_area(ws, 4, 4 + len(rep_stats), len(cols))
    for r in range(4, 4 + len(rep_stats)):
        ws.cell(row=r, column=3).number_format = '$#,##0'
        ws.cell(row=r, column=4).number_format = '$#,##0'
    add_table(ws, f"A3:E{3+len(rep_stats)}", "SalesRep_Table")

    for col_letter in "ABCDE":
        ws.column_dimensions[col_letter].width = 18

    # ============================================================
    # SHEET 5: monthly_trend (detail)
    # ============================================================
    ws = wb.create_sheet("monthly_trend")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "Monthly Revenue by Category (Detail)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    monthly_cat = df.groupby([df["Transaction_Date"].dt.month, "Category"])["Total_Revenue_USD"].sum().unstack(fill_value=0)
    monthly_cat.columns = [str(c) for c in monthly_cat.columns]
    monthly_cat.index = [month_names.get(int(m), str(m)) for m in monthly_cat.index]
    monthly_cat = monthly_cat.sort_index()
    monthly_cat["Total"] = monthly_cat.sum(axis=1)

    cols = ["Month"] + list(monthly_cat.columns)
    ws.append([])
    ws.append(cols)
    style_header_row(ws, 3, len(cols))
    for idx, row in monthly_cat.iterrows():
        ws.append([idx] + list(row))
    style_data_area(ws, 4, 3 + len(monthly_cat), len(cols))
    for r in range(4, 4 + len(monthly_cat)):
        for c in range(2, len(cols) + 1):
            ws.cell(row=r, column=c).number_format = '$#,##0'
    add_table(ws, f"A3:{get_column_letter(len(cols))}{3+len(monthly_cat)}", "MonthlyCat_Table")

    # Stacked bar chart
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "Monthly Revenue by Category"
    chart.y_axis.title = "Revenue ($)"
    data_ref = Reference(ws, min_col=2, min_row=3, max_row=3 + len(monthly_cat), max_col=len(monthly_cat.columns))
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=3 + len(monthly_cat))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 16
    chart.height = 9
    ws.add_chart(chart, "H3")

    for col_letter in "ABCDEFG":
        ws.column_dimensions[col_letter].width = 14

    # ============================================================
    # SHEET 6: price_tiers
    # ============================================================
    ws = wb.create_sheet("price_tiers")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    ws["A1"] = "Price Tier Analysis"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    tier_stats = df.groupby("Price_Tier").agg(
        Transactions=("Transaction_ID", "count"),
        Total_Revenue=("Total_Revenue_USD", "sum"),
        Avg_Unit_Price=("Unit_Price_USD", "mean"),
        Avg_Profit_Margin=("Profit_Margin", "mean"),
    ).reset_index()
    tier_stats["Avg_Unit_Price"] = tier_stats["Avg_Unit_Price"].round(2)
    tier_stats["Total_Revenue"] = tier_stats["Total_Revenue"].round(2)
    tier_stats["Avg_Profit_Margin"] = tier_stats["Avg_Profit_Margin"].apply(lambda x: f"{x*100:.1f}%")
    tier_order = ["Budget", "Mid", "Premium", "Flagship"]
    tier_stats["sort"] = tier_stats["Price_Tier"].map({t: i for i, t in enumerate(tier_order)})
    tier_stats = tier_stats.sort_values("sort").drop(columns="sort")

    cols = ["Price_Tier", "Transactions", "Total_Revenue", "Avg_Unit_Price", "Avg_Profit_Margin"]
    ws.append([])
    ws.append(cols)
    style_header_row(ws, 3, len(cols))
    for i, row in tier_stats.iterrows():
        ws.append(list(row))
    style_data_area(ws, 4, 3 + len(tier_stats), len(cols))
    for r in range(4, 4 + len(tier_stats)):
        ws.cell(row=r, column=3).number_format = '$#,##0'
        ws.cell(row=r, column=4).number_format = '$#,##0.00'
    add_table(ws, f"A3:E{3+len(tier_stats)}", "Tier_Table")

    for col_letter in "ABCDE":
        ws.column_dimensions[col_letter].width = 18

    # ============================================================
    # SHEET 7: ML_Model_Performance
    # ============================================================
    ws = wb.create_sheet("ML_Model_Performance")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "ML Model Performance — Revenue Prediction"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    ml_meta = [
        ("Target", results["target"]),
        ("Best Model", results["best_model"]),
        ("Best R² Score", results["best_r2_score"]),
        ("Test Size", f"{results['test_size']*100:.0f}%"),
        ("Train Rows", results["train_rows"]),
        ("Test Rows", results["test_rows"]),
        ("Features", results["n_features"]),
        ("Target Validation", results["verdict"]),
        ("Best Univariate Feature", results["best_univariate_feature"]),
        ("Preprocessing", "Median impute → IQR capping → StandardScaler"),
        ("Categoricals", "Rare group (<1% → OTHER) → OneHotEncoder"),
        ("Leakage Guard", "Fit on TRAIN only"),
    ]
    ws.append(["Parameter", "Value"])
    style_header_row(ws, 3, 2)
    for row in ml_meta:
        ws.append(list(row))
    style_data_area(ws, 4, 3 + len(ml_meta), 2)
    add_table(ws, "A3:B" + str(3 + len(ml_meta)), "MLMeta_Table")

    # Metrics table
    mstart = 4 + len(ml_meta) + 2
    ws.merge_cells(f"A{mstart}:E{mstart}")
    ws[f"A{mstart}"] = "Per-Model Metrics"
    ws[f"A{mstart}"].font = SUBTITLE_FONT
    ws[f"A{mstart}"].fill = PatternFill("solid", fgColor=LIGHT)
    mstart += 1
    ws.append([])
    ws.append(["Model", "R² Score", "MAE ($)", "RMSE ($)", "MAPE (%)"])
    style_header_row(ws, mstart, 5)
    r = mstart + 1
    for name, data in results["models"].items():
        ws.append([name, data["r2_score"], data["mae"], data["rmse"], data["mape_pct"]])
        r += 1
    style_data_area(ws, mstart + 1, r - 1, 5)
    for rr in range(mstart + 1, r):
        ws.cell(row=rr, column=2).number_format = '0.0000'
        ws.cell(row=rr, column=3).number_format = '$#,##0.00'
        ws.cell(row=rr, column=4).number_format = '$#,##0.00'
        ws.cell(row=rr, column=5).number_format = '0.0"%"'
    add_table(ws, f"A{mstart}:E{r-1}", "MLMetrics_Table")

    # Bar chart comparing R²
    chart = BarChart()
    chart.type = "col"
    chart.title = "R² Score by Model"
    chart.y_axis.title = "R²"
    data_ref = Reference(ws, min_col=2, min_row=mstart, max_row=r - 1)
    cats_ref = Reference(ws, min_col=1, min_row=mstart + 1, max_row=r - 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 12
    chart.height = 8
    ws.add_chart(chart, f"G{mstart}")

    for col_letter in "ABCDEFG":
        ws.column_dimensions[col_letter].width = 18

    # ============================================================
    # SHEET 8: ML_Features
    # ============================================================
    ws = wb.create_sheet("ML_Features")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"] = "Top 15 Features (GradientBoosting Importance)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    top_feats = results.get("top_features", {})
    sorted_feats = sorted(top_feats.items(), key=lambda x: x[1], reverse=True)[:15]

    ws.append([])
    ws.append(["Rank", "Feature", "Importance", "Normalized (%)"])
    style_header_row(ws, 3, 4)
    total_imp = sum(v for _, v in sorted_feats)
    for i, (feat, imp) in enumerate(sorted_feats, 1):
        ws.append([i, feat, imp, imp / total_imp * 100 if total_imp else 0])
    style_data_area(ws, 4, 3 + len(sorted_feats), 4)
    for rr in range(4, 4 + len(sorted_feats)):
        ws.cell(row=rr, column=3).number_format = '0.000000'
        ws.cell(row=rr, column=4).number_format = '0.0"%"'
    add_table(ws, "A3:D" + str(3 + len(sorted_feats)), "Feature_Table")

    # Conditional formatting on importance
    ws.conditional_formatting.add(
        f"C4:C{3+len(sorted_feats)}",
        ColorScaleRule(
            start_type="min", start_color="001E293B",
            end_type="max", end_color="0010B981",
        ),
    )

    # Bar chart
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Feature Importance (Top 15)"
    chart.x_axis.title = "Importance"
    data_ref = Reference(ws, min_col=3, min_row=3, max_row=3 + len(sorted_feats))
    cats_ref = Reference(ws, min_col=2, min_row=4, max_row=3 + len(sorted_feats))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 16
    chart.height = 10
    ws.add_chart(chart, "F3")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16

    # ============================================================
    # SAVE
    # ============================================================
    wb.save(EXCEL_REPORT_PATH)
    print(f"Excel report saved: {EXCEL_REPORT_PATH}")
    print(f"Size: {EXCEL_REPORT_PATH.stat().st_size:,} bytes")


if __name__ == "__main__":
    build_report()